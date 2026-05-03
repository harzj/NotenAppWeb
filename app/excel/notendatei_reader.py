"""
Reader for the special 'Notendatei' format used by colleagues.

Sheet structure (active / first sheet named 'Liste N'):
  A1:        Thema (topic/title of the exam)
  A3/C3:     'Kurs'  / Kursname  (class name)
  A4/C4:     'Termin' / date
  A6/C6:     'Punktesystem' / 15
  Row 9:     headers (Nr., Name, Vorname, Aufgabe, ...)
  Row 11:    task numbers (D+)
  Row 12:    'maximale Punkte', D+: max points per task
  Row 13+:   student data: Nr | Nachname | Vorname | pts... | Gesamt | Pct | Note6 | Note6text | Note15
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Optional

import msoffcrypto
import openpyxl
from openpyxl.workbook import Workbook


# ── helpers ───────────────────────────────────────────────────────────────────

def _open_wb(file_bytes: bytes, password: Optional[str] = None) -> Workbook:
    raw = io.BytesIO(file_bytes)
    try:
        ofd = msoffcrypto.OfficeFile(io.BytesIO(file_bytes))
        if ofd.is_encrypted():
            dec = io.BytesIO()
            ofd.load_key(password=password or "")
            ofd.decrypt(dec)
            dec.seek(0)
            return openpyxl.load_workbook(dec, data_only=True)
    except Exception:
        pass
    return openpyxl.load_workbook(raw, data_only=True)


def _str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _num(v) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def is_notendatei(file_bytes: bytes, password: Optional[str] = None) -> bool:
    """Quick check: is this a Notendatei (has 'Kurs' in A3 and task headers in row 12)?"""
    try:
        wb = _open_wb(file_bytes, password)
        ws = wb.active
        a3 = _str(ws.cell(3, 1).value).lower()
        a12 = _str(ws.cell(12, 1).value).lower()
        return "kurs" in a3 and "maximale punkte" in a12
    except Exception:
        return False


# ── public API ────────────────────────────────────────────────────────────────

def read_notendatei(file_bytes: bytes, password: Optional[str] = None) -> dict:
    """
    Parse a Notendatei Excel file and return a partial gradebook dict with:
      - klasse, thema, datum, aufgaben, schueler
    This returns a single LN dict ready to be appended to leistungsnachweise.
    """
    wb = _open_wb(file_bytes, password)

    # Use the first sheet (usually 'Liste 1')
    ws = wb.active

    # ── Meta info ─────────────────────────────────────────────────────────────
    thema = _str(ws.cell(1, 1).value)          # A1: Thema
    klasse = _str(ws.cell(3, 3).value)         # C3: Kurs
    termin_raw = ws.cell(4, 3).value           # C4: Termin (datetime or str)
    datum = ""
    if isinstance(termin_raw, datetime):
        datum = termin_raw.strftime("%Y-%m-%d")
    elif termin_raw is not None:
        datum = _str(termin_raw)

    # ── Task columns (row 11 = task numbers, row 12 = max points) ────────────
    # Task data starts at column D (col 4).
    # Only include columns where row 11 has a numeric task number (1, 2, 3...).
    # Summary columns like "Erreichte Punktzahl" (no number in row 11) are excluded.
    START_COL = 4

    task_cols: list[int] = []    # 1-based column indices
    task_maxpts: list[float] = []
    task_labels: list[str] = []

    max_col = ws.max_column or 30
    for c in range(START_COL, max_col + 1):
        num_label = ws.cell(11, c).value
        # Stop when there's no numeric task label in row 11
        if not isinstance(num_label, (int, float)):
            break
        max_pt = ws.cell(12, c).value
        mp = _num(max_pt)
        if mp is not None and mp > 0:
            task_cols.append(c)
            task_maxpts.append(mp)
            task_labels.append(str(int(num_label)))

    # Build aufgaben list
    aufgaben = []
    for i in range(len(task_cols)):
        aufgaben.append({
            "label": task_labels[i],
            "afb": "",
            "max_punkte": task_maxpts[i],
        })

    # ── Student rows (row 13+) ────────────────────────────────────────────────
    # Stop as soon as col A is not a positive sequential integer
    # (the Notenspiegel section has text like "Gültigkeit:", "Notenspiegel:" etc.)
    schueler = []
    for r in range(13, ws.max_row + 1):
        nr_val = ws.cell(r, 1).value
        if nr_val is None:
            continue
        try:
            nr = int(float(nr_val))
            if nr <= 0:
                break
        except (ValueError, TypeError):
            break  # Hit non-data section

        nachname_raw = ws.cell(r, 2).value   # col B
        vorname_raw  = ws.cell(r, 3).value   # col C

        # Skip rows with no name (empty student slots)
        nachname = _str(nachname_raw)
        vorname  = _str(vorname_raw)
        if not nachname and not vorname:
            continue

        # Full name in "Nachname, Vorname" convention matching stammdaten
        full_name = f"{nachname}, {vorname}" if vorname else nachname

        # Read points for each task column
        punkte: list[Optional[float]] = []
        for col in task_cols:
            val = ws.cell(r, col).value
            p = _num(val)
            punkte.append(p)

        schueler.append({
            "name": full_name,
            "punkte": punkte,
            "note_15": None,
            "note_6": None,
            "ignoriert": False,
        })

    return {
        "klasse": klasse,
        "thema": thema,
        "datum": datum,
        "aufgaben": aufgaben,
        "aufgaben_tree": [],
        "schueler": schueler,
    }
