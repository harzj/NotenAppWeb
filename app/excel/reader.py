"""
Read a (possibly password-protected) xlsx file into a structured Python dict.
All processing happens in-memory (BytesIO); the file never touches disk.
"""
from __future__ import annotations

import io
import re
from datetime import date as date_cls
from typing import Any

import msoffcrypto
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.excel import schema as S


class ExcelReadError(Exception):
    pass


def schuljahr_from_date(d: date_cls | None = None) -> str:
    """Return school-year string like '2526' for 2025/26.
    Cutoff: before July 21 → previous school year, else current."""
    if d is None:
        d = date_cls.today()
    cutoff = date_cls(d.year, 7, 21)
    start = d.year - 1 if d < cutoff else d.year
    return f"{start % 100:02d}{(start + 1) % 100:02d}"


# ── Public entry point ────────────────────────────────────────────────────────

def load_gradebook(file_bytes: bytes, password: str | None = None) -> dict:
    """
    Decrypt (if needed) and parse an xlsx grade file.

    Returns a dict with keys:
        klasse: str
        fach: str
        schuljahr: str
        stammdaten: list[dict]
        leistungsnachweise: list[dict]
        uebersicht_hj1: dict | None
        uebersicht_hj2: dict | None
        uebersicht_jahr: dict | None
        mdl_noten: dict
        sl_noten_actual: dict
        hj_noten: dict
        schuljahr_noten_actual: dict
        sl_gewichtung: dict | None
    """
    wb = _open_workbook(file_bytes, password)
    fach, schuljahr = _read_metadata(wb)
    result = {
        "klasse": _read_class_name(wb),
        "fach": fach,
        "schuljahr": schuljahr,
        "stammdaten": _read_stammdaten(wb),
        "leistungsnachweise": _read_all_ln(wb),
        "uebersicht_hj1": _read_uebersicht(wb, S.SHEET_UEBERSICHT_HJ1),
        "uebersicht_hj2": _read_uebersicht(wb, S.SHEET_UEBERSICHT_HJ2),
        "uebersicht_jahr": _read_uebersicht(wb, S.SHEET_UEBERSICHT_JAHR),
    }
    noten_zusatz = _read_noten_zusatz(wb)
    result.update(noten_zusatz)
    settings = _read_einstellungen(wb)
    if settings:
        # Separate sl_gewichtung from kurs settings
        sl_gw = {k: v for k, v in settings.items() if k in S.ES_GEWICHTUNG_KEYS}
        if sl_gw:
            result["sl_gewichtung"] = sl_gw
        # Kurs settings
        modus = settings.get("modus", "klasse")
        result["modus"] = modus if modus in ("klasse", "kurs") else "klasse"
        kurs_typ = settings.get("kurs_typ", "")
        if kurs_typ:
            result["kurs_typ"] = kurs_typ
        stunden_raw = settings.get("kurs_stunden", "")
        if stunden_raw:
            try:
                result["kurs_stunden"] = int(float(stunden_raw))
            except (TypeError, ValueError):
                pass
        kurs_gln = settings.get("kurs_gln_pct", "")
        kurs_mdl = settings.get("kurs_mdl_pct", "")
        if kurs_gln or kurs_mdl:
            result.setdefault("kurs_gewichtung", {})
            try:
                result["kurs_gewichtung"]["hj_gln_pct"] = float(kurs_gln)
            except (TypeError, ValueError):
                pass
            try:
                result["kurs_gewichtung"]["hj_mdl_pct"] = float(kurs_mdl)
            except (TypeError, ValueError):
                pass
    else:
        result.setdefault("modus", "klasse")
    return result


# ── Workbook opening ─────────────────────────────────────────────────────────

def _open_workbook(file_bytes: bytes, password: str | None) -> Workbook:
    buf = io.BytesIO(file_bytes)
    try:
        office_file = msoffcrypto.OfficeFile(buf)
        if office_file.is_encrypted():
            if not password:
                raise ExcelReadError("Die Datei ist passwortgeschützt. Bitte Passwort angeben.")
            decrypted = io.BytesIO()
            office_file.load_key(password=password)
            office_file.decrypt(decrypted)
            decrypted.seek(0)
            source = decrypted
        else:
            buf.seek(0)
            source = buf
    except msoffcrypto.exceptions.InvalidKeyError:
        raise ExcelReadError("Falsches Passwort.")
    except Exception as exc:
        raise ExcelReadError(f"Fehler beim Öffnen der Datei: {exc}") from exc

    try:
        wb = openpyxl.load_workbook(source, data_only=False)
    except Exception as exc:
        raise ExcelReadError(f"Ungültiges Excel-Format: {exc}") from exc
    return wb


# ── Class name ────────────────────────────────────────────────────────────────

def _read_class_name(wb: Workbook) -> str:
    if S.SHEET_STAMMDATEN not in wb.sheetnames:
        return ""
    ws: Worksheet = wb[S.SHEET_STAMMDATEN]
    row = list(ws.iter_rows(
        min_row=S.SD_INFO_ROW, max_row=S.SD_INFO_ROW, values_only=True
    ))[0]
    return _str(row, S.SD_CLASS_NAME_VALUE_COL - 1)


def _read_metadata(wb: Workbook) -> tuple[str, str]:
    """Returns (fach, schuljahr) from Stammdaten info row.
    Falls back to auto-detected school year if not stored."""
    if S.SHEET_STAMMDATEN not in wb.sheetnames:
        return "", schuljahr_from_date()
    ws: Worksheet = wb[S.SHEET_STAMMDATEN]
    row = list(ws.iter_rows(
        min_row=S.SD_INFO_ROW, max_row=S.SD_INFO_ROW, values_only=True
    ))[0]
    fach = _str(row, S.SD_FACH_VALUE_COL - 1)
    sj = _str(row, S.SD_SJ_VALUE_COL - 1)
    if not sj:
        sj = schuljahr_from_date()
    return fach, sj


# ── Stammdaten ────────────────────────────────────────────────────────────────

def _read_stammdaten(wb: Workbook) -> list[dict]:
    if S.SHEET_STAMMDATEN not in wb.sheetnames:
        return []
    ws: Worksheet = wb[S.SHEET_STAMMDATEN]
    students = []
    for row in ws.iter_rows(min_row=S.SD_DATA_START_ROW, values_only=True):
        if not any(row):
            continue
        students.append({
            "nachname": _str(row, S.SD_COL_NACHNAME - 1),
            "vorname":  _str(row, S.SD_COL_VORNAME  - 1),
            "status":   _str(row, S.SD_COL_STATUS   - 1) or S.SD_STATUS_AKTIV,
            "austritt": _date_str(row, S.SD_COL_AUSTRITT - 1),
            "abgang_nach_hj": _str(row, S.SD_COL_ABGANG_HJ - 1) or None,
        })
    return students


# ── Leistungsnachweise ────────────────────────────────────────────────────────

def _read_all_ln(wb: Workbook) -> list[dict]:
    sd_ws = wb[S.SHEET_STAMMDATEN] if S.SHEET_STAMMDATEN in wb.sheetnames else None
    result = []
    for name in wb.sheetnames:
        if name.startswith(S.LN_SHEET_PREFIX):
            result.append(_read_ln_sheet(wb[name], name, stammdaten_ws=sd_ws))
    return result


def _read_ln_sheet(ws: Worksheet, sheet_name: str, stammdaten_ws: Worksheet | None = None) -> dict:
    ln_name = sheet_name[len(S.LN_SHEET_PREFIX):]

    # ── Detect format: v2 has a metadata row in row 1 ──────────────────────
    first_cell = ws.cell(1, 1).value
    is_v2 = isinstance(first_cell, str) and first_cell.strip() == S.LN_META_TYP_LABEL

    if is_v2:
        row_header    = S.LN_ROW_HEADER
        row_afb       = S.LN_ROW_AFB
        row_max       = S.LN_ROW_MAX
        data_start    = S.LN_DATA_START_ROW
        ln_typ        = _str(list(ws.iter_rows(min_row=S.LN_ROW_META, max_row=S.LN_ROW_META,
                                               values_only=True))[0], S.LN_META_TYP_VAL - 1)
        _meta_row     = list(ws.iter_rows(min_row=S.LN_ROW_META, max_row=S.LN_ROW_META,
                                          values_only=True))[0]
        hj            = _str(_meta_row, S.LN_META_HJ_VAL - 1) or None
        sl_zuordnung  = _str(_meta_row, S.LN_META_SL_VAL - 1) or None
        gln_slot      = _str(_meta_row, S.LN_META_GSLOT_VAL - 1) or None
        nachtermin_von = _str(_meta_row, S.LN_META_NT_VAL - 1) or None
        runden_raw    = _str(_meta_row, S.LN_META_RUNDEN_VAL - 1)
        noten_runden  = (runden_raw != "0")  # default True unless explicitly "0"
    else:
        # Old format: no meta row – rows are shifted by -1
        row_header    = 1
        row_afb       = 2
        row_max       = 3
        data_start    = 4
        ln_typ        = None
        hj            = None
        sl_zuordnung  = None
        gln_slot      = None
        nachtermin_von = None
        noten_runden  = True

    header_row = list(ws.iter_rows(min_row=row_header, max_row=row_header, values_only=True))[0]
    afb_row    = list(ws.iter_rows(min_row=row_afb,    max_row=row_afb,    values_only=True))[0]
    max_row    = list(ws.iter_rows(min_row=row_max,    max_row=row_max,    values_only=True))[0]

    # Detect task columns: between col 2 and "Gesamt" column
    task_cols: list[dict] = []
    gesamt_col_idx = None
    ignoriert_col_idx = None
    for i, val in enumerate(header_row):
        if i == 0:
            continue  # name column
        if isinstance(val, str) and val.strip() == S.LN_HEADER_GESAMT:
            gesamt_col_idx = i
            break
        task_cols.append({
            "label": _str(header_row, i) or f"Aufgabe{i}",
            "afb": _str(afb_row, i),
            "max_punkte": _num(max_row, i),
        })

    # Find Ignoriert column (after Note 1-6)
    if gesamt_col_idx is not None:
        for i in range(gesamt_col_idx + 1, len(header_row)):
            v = header_row[i]
            if isinstance(v, str) and v.strip() == S.LN_HEADER_IGNORIERT:
                ignoriert_col_idx = i
                break

    # Read student rows
    students = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        raw_name = row[0]
        if not raw_name:
            continue
        if isinstance(raw_name, str) and raw_name.startswith("=") and stammdaten_ws is not None:
            m = re.search(r'!A(\d+)', raw_name)
            if m:
                sd_row = int(m.group(1))
                nn = stammdaten_ws.cell(sd_row, S.SD_COL_NACHNAME).value or ""
                vn = stammdaten_ws.cell(sd_row, S.SD_COL_VORNAME).value or ""
                student_name = f"{nn}, {vn}".strip(", ") if (nn or vn) else _str(row, 0)
            else:
                student_name = _str(row, 0)
        else:
            student_name = _str(row, 0)

        punkte = []
        for t_idx in range(len(task_cols)):
            col_i = S.LN_COL_TASKS_START - 1 + t_idx
            punkte.append(_num(row, col_i))

        note15 = None
        note6  = None
        if gesamt_col_idx is not None:
            note15 = _num(row, gesamt_col_idx + 1)
            note6  = _num(row, gesamt_col_idx + 2)

        # Fallback: recompute note_15 from punkte when the cell contains a
        # formula string (data_only=False) that _num() cannot parse.
        if note15 is None and any(p is not None for p in punkte):
            max_total = sum(t.get("max_punkte") or 0 for t in task_cols)
            if max_total > 0:
                total = sum(p for p in punkte if p is not None)
                note15 = S.percent_to_note15(total, max_total, runden=noten_runden)
                note6  = S.note15_to_note6(note15)

        ignoriert = False
        if ignoriert_col_idx is not None:
            val = row[ignoriert_col_idx] if ignoriert_col_idx < len(row) else None
            ignoriert = bool(val) if val is not None else False

        students.append({
            "name":      student_name,
            "punkte":    punkte,
            "note_15":   int(note15) if note15 is not None else None,
            "note_6":    int(note6)  if note6  is not None else None,
            "ignoriert": ignoriert,
        })

    # Reconstruct aufgaben_tree from labels (detect hierarchy by dot-notation)
    aufgaben_tree = _rebuild_tree_from_labels(task_cols)

    return {
        "sheet_name":    sheet_name,
        "name":          ln_name,
        "ln_typ":        ln_typ or "KLN",
        "hj":            hj,
        "sl_zuordnung":  sl_zuordnung,
        "gln_slot":      gln_slot,
        "nachtermin_von": nachtermin_von,
        "noten_runden":  noten_runden,
        "aufgaben":      task_cols,
        "aufgaben_tree": aufgaben_tree,
        "schueler":      students,
    }


def _rebuild_tree_from_labels(task_cols: list[dict]) -> list[dict]:
    """
    Reconstruct a hierarchical aufgaben_tree from flat task_cols whose labels
    follow the standard dot-notation (e.g. "1", "1.a", "1.b", "2", "2.1", "2.1.a").

    Falls back to a flat tree (each task as top-level leaf) if the labels do not
    match the expected pattern.
    """
    import re as _re

    def _make_leaf(task: dict, node_id: str) -> dict:
        label = task.get("label", "")
        # Detect numbering style of children from label suffix
        style = "123"
        return {
            "id": node_id,
            "label": label,
            "custom_label": bool(label),
            "afb": task.get("afb", ""),
            "max_punkte": task.get("max_punkte"),
            "numbering_style": style,
            "children": [],
        }

    # Check if any label contains a dot → hierarchical
    labels = [t.get("label", "") for t in task_cols]
    has_hierarchy = any("." in lbl for lbl in labels)

    if not has_hierarchy:
        # Plain flat tree
        return [_make_leaf(t, f"t{i}") for i, t in enumerate(task_cols)]

    # Build tree by grouping top-level parts
    # A label like "1.a" → parent "1", child "a"
    # A label like "1.1.a" → parent "1", child "1", grand-child "a"
    root_nodes: dict[str, dict] = {}
    root_order: list[str] = []
    node_counter = [0]

    def _nid():
        node_counter[0] += 1
        return f"t{node_counter[0]}"

    for task in task_cols:
        label = task.get("label", "")
        parts = label.split(".")
        top = parts[0]
        if top not in root_nodes:
            root_nodes[top] = {
                "id": _nid(),
                "label": top,
                "custom_label": True,
                "afb": "",
                "max_punkte": None,
                "numbering_style": "123",
                "children": [],
            }
            root_order.append(top)

        if len(parts) == 1:
            # Top-level leaf – update with actual task data
            root_nodes[top]["afb"] = task.get("afb", "")
            root_nodes[top]["max_punkte"] = task.get("max_punkte")
        elif len(parts) == 2:
            child_label = parts[1]
            # Detect style
            if _re.match(r'^[a-z]$', child_label):
                root_nodes[top]["numbering_style"] = "abc"
            child = _make_leaf({"label": child_label, "afb": task.get("afb", ""),
                                 "max_punkte": task.get("max_punkte")}, _nid())
            root_nodes[top]["children"].append(child)
        else:
            # 3-level: parts[0].parts[1].parts[2]
            mid_label = parts[1]
            # Find or create mid node
            mid = None
            for c in root_nodes[top]["children"]:
                if c["label"] == mid_label:
                    mid = c
                    break
            if mid is None:
                mid = {
                    "id": _nid(),
                    "label": mid_label,
                    "custom_label": True,
                    "afb": "",
                    "max_punkte": None,
                    "numbering_style": "123",
                    "children": [],
                }
                root_nodes[top]["children"].append(mid)
            child_label = parts[2]
            if _re.match(r'^[a-z]$', child_label):
                mid["numbering_style"] = "abc"
            grandchild = _make_leaf({"label": child_label, "afb": task.get("afb", ""),
                                      "max_punkte": task.get("max_punkte")}, _nid())
            mid["children"].append(grandchild)

    return [root_nodes[k] for k in root_order]


def _read_noten_zusatz(wb: Workbook) -> dict:
    """Read per-student MDL/SL/HJ/SJ notes from the hidden Noten_Zusatz sheet."""
    empty = {
        "mdl_noten": {},
        "sl_noten_actual": {},
        "hj_noten": {},
        "schuljahr_noten_actual": {},
        "mdl_noten_kurs": {},
    }
    if S.SHEET_NOTEN_ZUSATZ not in wb.sheetnames:
        return empty

    ws: Worksheet = wb[S.SHEET_NOTEN_ZUSATZ]
    mdl: dict = {}
    sl_act: dict = {}
    hj_act: dict = {}
    sj_act: dict = {}
    mdl_kurs: dict = {}

    for row in ws.iter_rows(min_row=S.NZ_DATA_START, values_only=True):
        name = _str(row, S.NZ_COL_NAME - 1)
        if not name:
            continue
        def _intn(row, col):
            v = _num(row, col - 1)
            return int(v) if v is not None else None

        mdl[name] = {
            "SL1": _intn(row, S.NZ_COL_MDL_SL1),
            "SL2": _intn(row, S.NZ_COL_MDL_SL2),
            "SL3": _intn(row, S.NZ_COL_MDL_SL3),
            "SL4": _intn(row, S.NZ_COL_MDL_SL4),
        }
        sl_act[name] = {
            "SL1": _intn(row, S.NZ_COL_SL_ACT_SL1),
            "SL2": _intn(row, S.NZ_COL_SL_ACT_SL2),
            "SL3": _intn(row, S.NZ_COL_SL_ACT_SL3),
            "SL4": _intn(row, S.NZ_COL_SL_ACT_SL4),
        }
        hj_act[name] = {
            "HJ1": _intn(row, S.NZ_COL_HJ_ACT_HJ1),
            "HJ2": _intn(row, S.NZ_COL_HJ_ACT_HJ2),
            "HJ3": _intn(row, S.NZ_COL_KURS_HJ_ACT_HJ3),
            "HJ4": _intn(row, S.NZ_COL_KURS_HJ_ACT_HJ4),
        }
        sj_val = _intn(row, S.NZ_COL_SJ_ACT)
        if sj_val is not None:
            sj_act[name] = sj_val
        mdl_kurs[name] = {
            "HJ1_mdl1": _intn(row, S.NZ_COL_KURS_MDL_HJ1_1),
            "HJ1_mdl2": _intn(row, S.NZ_COL_KURS_MDL_HJ1_2),
            "HJ2_mdl1": _intn(row, S.NZ_COL_KURS_MDL_HJ2_1),
            "HJ2_mdl2": _intn(row, S.NZ_COL_KURS_MDL_HJ2_2),
            "HJ3_mdl1": _intn(row, S.NZ_COL_KURS_MDL_HJ3_1),
            "HJ3_mdl2": _intn(row, S.NZ_COL_KURS_MDL_HJ3_2),
            "HJ4_mdl1": _intn(row, S.NZ_COL_KURS_MDL_HJ4_1),
            "HJ4_mdl2": _intn(row, S.NZ_COL_KURS_MDL_HJ4_2),
        }

    return {
        "mdl_noten": mdl,
        "sl_noten_actual": sl_act,
        "hj_noten": hj_act,
        "schuljahr_noten_actual": sj_act,
        "mdl_noten_kurs": mdl_kurs,
    }


def _read_einstellungen(wb: Workbook) -> dict | None:
    """Read sl_gewichtung and Kurs settings from hidden Einstellungen sheet."""
    if S.SHEET_EINSTELLUNGEN not in wb.sheetnames:
        return None
    ws: Worksheet = wb[S.SHEET_EINSTELLUNGEN]
    result: dict = {}
    for row in ws.iter_rows(min_row=S.ES_DATA_START, values_only=True):
        key = _str(row, S.ES_COL_KEY - 1)
        val_raw = row[S.ES_COL_VALUE - 1] if len(row) > S.ES_COL_VALUE - 1 else None
        if key in S.ES_GEWICHTUNG_KEYS:
            val = _num(row, S.ES_COL_VALUE - 1)
            if val is not None:
                result[key] = val
        elif key in S.ES_KURS_KEYS:
            result[key] = str(val_raw).strip() if val_raw is not None else ""
    return result if result else None


# ── Übersichten ───────────────────────────────────────────────────────────────

def _read_uebersicht(wb: Workbook, sheet_name: str) -> dict | None:
    if sheet_name not in wb.sheetnames:
        return None
    ws: Worksheet = wb[sheet_name]
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    columns = [v for v in header_row if v is not None]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        rows.append(list(row[: len(columns)]))
    return {"columns": columns, "rows": rows}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _str(row: tuple, idx: int) -> str:
    try:
        v = row[idx]
        return str(v).strip() if v is not None else ""
    except IndexError:
        return ""


def _num(row: tuple, idx: int) -> float | None:
    try:
        v = row[idx]
        if v is None:
            return None
        return float(v)
    except (IndexError, TypeError, ValueError):
        return None


def _date_str(row: tuple, idx: int) -> str:
    try:
        v = row[idx]
        if v is None:
            return ""
        if hasattr(v, "strftime"):
            return v.strftime("%d.%m.%Y")
        return str(v).strip()
    except IndexError:
        return ""
