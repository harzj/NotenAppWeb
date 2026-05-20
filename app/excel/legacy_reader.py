"""
Legacy Excel gradebook reader (old format, pre-NotenAppWeb).

Sheet structure found in old files:
  GLN1, GLN 1, GLN2…   – Große Leistungsnachweise
  KLN 1, KLN 2…        – Kleine LN sets, each containing multiple sub-tests (HÜ)
  Noten                 – HJ1 summary: MDL1/MDL2, HJ1 Zeugnisnote, SL1/SL2 actuals
  Noten (2)             – full-year summary: MDL3/MDL4, Ganzjahresnote, SL3/SL4 actuals
"""
from __future__ import annotations

import io
import re
from typing import Optional

import msoffcrypto
import openpyxl
from openpyxl.workbook import Workbook

from app.excel import schema as S


# ── helpers ───────────────────────────────────────────────────────────────────

def _open_wb(file_bytes: bytes, password: Optional[str] = None) -> Workbook:
    raw = io.BytesIO(file_bytes)
    try:
        ofd = msoffcrypto.OfficeFile(io.BytesIO(file_bytes))
        if ofd.is_encrypted():
            dec = io.BytesIO()
            ofd.load_key(password=password or "")
            ofd.decrypt(dec)
            dec.seek(0)
            return openpyxl.load_workbook(dec, data_only=True)
    except Exception:
        pass
    return openpyxl.load_workbook(raw, data_only=True)


def _cell_str(ws, row: int, col: int) -> Optional[str]:
    v = ws.cell(row, col).value
    if v is None:
        return None
    s = str(v).strip()
    return s if s and not s.startswith("#") else None


def _cell_num(ws, row: int, col: int) -> Optional[float]:
    v = ws.cell(row, col).value
    if v is None:
        return None
    try:
        f = float(v)
        return f if f >= 0 else None
    except (ValueError, TypeError):
        return None


def _int_or_none(v) -> Optional[int]:
    if v is None:
        return None
    try:
        return round(float(v))
    except (ValueError, TypeError):
        return None


# ── student extraction ────────────────────────────────────────────────────────

_SKIP_NAMES = {"max punkte", "teilaufgabe", "gewicht"}
_SKIP_KLN_COLS = {"notentabelle", "mw", "mittelwert", "gesamt", "note", "note(15)", "note(6)", "summe"}


def _get_students(wb: Workbook) -> list[str]:
    """Extract ordered student name list from Noten or first available sheet."""
    for sheet_name in ["Noten", "KLN 1", "KLN1", "GLN1", "GLN 1"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        names = []
        for r in range(4, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if not v:
                continue
            name = str(v).strip()
            if not name or name.startswith("#") or name.lower() in _SKIP_NAMES:
                continue
            names.append(name)
        if names:
            return names
    return []


# ── KLN sheet scanning ────────────────────────────────────────────────────────

def _scan_kln_sheet(ws) -> list[dict]:
    """Return list of importable HÜ entries {col, name, max_pts}."""
    hue_list = []
    max_col = ws.max_column or 20
    row1 = [ws.cell(1, c).value for c in range(1, max_col + 1)]
    row3 = [ws.cell(3, c).value for c in range(1, max_col + 1)]

    for i, name in enumerate(row1):
        if not isinstance(name, str):
            continue
        name = name.strip()
        if not name or name.lower() in _SKIP_KLN_COLS:
            continue
        col = i + 1  # 1-based
        raw_max = row3[i] if i < len(row3) else None
        try:
            max_pts = float(raw_max) if raw_max is not None else 0.0
        except (ValueError, TypeError):
            max_pts = 0.0
        if max_pts > 0:
            hue_list.append({"col": col, "name": name, "max_pts": round(max_pts)})
    return hue_list


# ── public: probe file ────────────────────────────────────────────────────────

def probe_legacy_file(file_bytes: bytes, password: Optional[str] = None) -> dict:
    """
    Open the legacy file and return a JSON-serialisable structure description
    for use in the import wizard.
    """
    wb = _open_wb(file_bytes, password)

    # ── Detect Oberstufen-Kurs format ─────────────────────────────────────────
    # Criterion: sheets GLN1 and GLN2 exist AND no KLN sheets
    gln_numbers = []
    for name in wb.sheetnames:
        m = re.match(r"^GLN\s*(\d+)$", name, re.I)
        if m:
            gln_numbers.append((int(m.group(1)), name))
    gln_numbers.sort()
    has_kln = any(re.match(r"^KLN\s*\d+$", n, re.I) for n in wb.sheetnames)
    is_oberstufe = len(gln_numbers) >= 2 and not has_kln

    if is_oberstufe:
        students = _get_students_oberstufe(wb)
        gln_sheets = [name for _, name in gln_numbers]
        # Detect Zeugnisnoten sheets
        zeugnisnoten_sheets = [
            n for n in wb.sheetnames
            if re.match(r"^Zeugnisnoten(\s*\(\d+\))?$", n, re.I)
        ]
        return {
            "students":           students,
            "student_count":      len(students),
            "is_oberstufe":       True,
            "kln_sheets":         [],
            "gln_sheets":         gln_sheets,
            "gln_auto_slots":     _gln_auto_slots(gln_sheets),
            "zeugnisnoten_sheets": zeugnisnoten_sheets,
            "has_noten":          False,
            "has_noten2":         False,
        }

    # ── Standard (Klasse) format ──────────────────────────────────────────────
    students = _get_students(wb)
    kln_sheets: list[dict] = []
    gln_sheets_std: list[str] = []

    for name in wb.sheetnames:
        if re.match(r"^KLN\s*\d+$", name, re.I):
            hue_list = _scan_kln_sheet(wb[name])
            if hue_list:
                kln_sheets.append({"sheet": name, "hue_list": hue_list})
        elif re.match(r"^GLN\s*\d+$", name, re.I):
            gln_sheets_std.append(name)

    return {
        "students":      students,
        "student_count": len(students),
        "is_oberstufe":  False,
        "kln_sheets":    kln_sheets,
        "gln_sheets":    gln_sheets_std,
        "has_noten":     "Noten" in wb.sheetnames,
        "has_noten2":    "Noten (2)" in wb.sheetnames,
    }


# ── Auto-slot mapping for Oberstufe GLN sheets ────────────────────────────────

_GLN_SLOT_MAP = {
    1: ("GLN1", "HJ1"), 2: ("GLN2", "HJ1"),
    3: ("GLN3", "HJ2"), 4: ("GLN4", "HJ2"),
    5: ("GLN5", "HJ3"), 6: ("GLN6", "HJ3"),
    7: ("GLN7", "HJ4"), 8: ("GLN8", "HJ4"),
}


def _gln_auto_slots(gln_sheets: list[str]) -> dict:
    """Return {sheet_name: {slot, hj}} for auto-assignment."""
    result = {}
    for name in gln_sheets:
        m = re.match(r"^GLN\s*(\d+)$", name, re.I)
        if m:
            num = int(m.group(1))
            slot, hj = _GLN_SLOT_MAP.get(num, (f"GLN{num}", "HJ1"))
            result[name] = {"slot": slot, "hj": hj}
    return result


# ── Oberstufe: student extraction from Zeugnisnoten ──────────────────────────

def _get_students_oberstufe(wb: Workbook) -> list[str]:
    """Extract student names from the first available Zeugnisnoten sheet (col A, row 6+)."""
    candidates = [n for n in wb.sheetnames if re.match(r"^Zeugnisnoten(\s*\(\d+\))?$", n, re.I)]
    # Also fall back to GLN1 if no Zeugnisnoten sheet found
    if not candidates:
        candidates = [n for n in wb.sheetnames if re.match(r"^GLN\s*1$", n, re.I)]
    for sheet_name in candidates:
        ws = wb[sheet_name]
        names = []
        for r in range(6, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if not v:
                continue
            name = str(v).strip()
            if not name or name.startswith("#") or name.lower() in _SKIP_NAMES:
                continue
            names.append(name)
        if names:
            return names
    return []


# ── Oberstufe: GLN sheet reader (new Oberstufe format) ───────────────────────

def _read_oberstufe_gln(wb: Workbook, sheet_name: str, students: list[str],
                         gln_slot: str, hj: str) -> Optional[dict]:
    """
    Read an Oberstufen-GLN sheet.

    Layout:
      Row 4: some cells; "Gesamt" label appears in the Summen-column (e.g. col 14)
      Row 5: Max-Punkte per Aufgabe in cols B...(gesamt_col-1), Gesamt in gesamt_col
      Row 6+: Col A = Schülername, Col B..gesamt_col-1 = Punkte,
              Col gesamt_col = Gesamtpunkte (formula, may be None),
              Col gesamt_col+1 = Note 15P
    """
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]

    # Find "Gesamt" entry in row 4
    max_col = ws.max_column or 30
    gesamt_col: Optional[int] = None
    for c in range(1, max_col + 1):
        v = ws.cell(4, c).value
        if isinstance(v, str) and "gesamt" in v.strip().lower():
            gesamt_col = c
            break
    if gesamt_col is None or gesamt_col <= 2:
        # Fallback: use old-style "Gesamt" in row 2
        for c in range(1, max_col + 1):
            v = ws.cell(2, c).value
            if isinstance(v, str) and "gesamt" in v.strip().lower():
                gesamt_col = c
                break

    if gesamt_col is None:
        gesamt_col = max_col  # last resort

    # Row 5: max points for tasks in cols B..gesamt_col-1
    aufgaben = []
    for c in range(2, gesamt_col):
        raw = ws.cell(5, c).value
        try:
            mp = float(raw) if raw is not None else None
        except (ValueError, TypeError):
            mp = None
        if mp is not None and mp > 0:
            aufgaben.append({"label": str(c - 1), "afb": "", "max_punkte": mp})

    if not aufgaben:
        # Fallback: treat as single-task GLN with total from gesamt_col row 5
        raw = ws.cell(5, gesamt_col).value if gesamt_col else None
        try:
            mp = float(raw) if raw is not None else 100.0
        except (ValueError, TypeError):
            mp = 100.0
        aufgaben = [{"label": "Gesamt", "afb": "", "max_punkte": mp}]

    # Row 6+: student data
    schueler = []
    for r in range(6, ws.max_row + 1):
        raw_name = ws.cell(r, 1).value
        if not raw_name:
            continue
        name = str(raw_name).strip()
        if not name or name.startswith("#") or name.lower() in _SKIP_NAMES:
            continue

        punkte: list[Optional[float]] = []
        if len(aufgaben) == 1 and aufgaben[0]["label"] == "Gesamt":
            # Single-task fallback: read from gesamt_col
            raw_pts = ws.cell(r, gesamt_col).value if gesamt_col else None
            try:
                pts = round(float(raw_pts), 2) if raw_pts is not None else None
            except (ValueError, TypeError):
                pts = None
            punkte = [pts]
        else:
            for c in range(2, gesamt_col):
                col_idx = c - 2
                if col_idx >= len(aufgaben):
                    break
                raw_pts = ws.cell(r, c).value
                try:
                    pts = round(float(raw_pts), 2) if raw_pts is not None else None
                except (ValueError, TypeError):
                    pts = None
                punkte.append(pts)

        # Note 15P is in col gesamt_col+1
        note15: Optional[int] = None
        if gesamt_col:
            note15 = _int_or_none(ws.cell(r, gesamt_col + 1).value)

        schueler.append({
            "name": name,
            "punkte": punkte,
            "note_15": note15,
            "note_6": None,
            "ignoriert": False,
        })

    ln_name = sheet_name.replace(" ", "")
    return {
        "sheet_name":    f"LN_{ln_name}",
        "name":          ln_name,
        "ln_typ":        "GLN",
        "hj":            hj,
        "gln_slot":      gln_slot,
        "sl_zuordnung":  None,
        "nachtermin_von": None,
        "noten_runden":  False,
        "thema":         "",
        "datum":         "",
        "aufgaben":      aufgaben,
        "aufgaben_tree": [],
        "schueler":      schueler,
    }


# ── Oberstufe: Zeugnisnoten sheet reader ─────────────────────────────────────

_ZEUGNIS_HJ_MAP = {
    "Zeugnisnoten": "HJ1",
    "Zeugnisnoten (2)": "HJ2",
    "Zeugnisnoten (3)": "HJ3",
    "Zeugnisnoten (4)": "HJ4",
}


def _read_zeugnisnoten(ws, students: list[str], hj_key: str) -> dict:
    """
    Extract from a Zeugnisnoten sheet:
      - Row 5: B5=Gewicht GLN1, C5=Gewicht GLN2, D5=Gewicht MDL1, E5=Gewicht MDL2
      - Row 4: "Note" label appears in some column (e.g. H4) → note_col
      - Row 6+: Col A = Schülername, Note from note_col
    Returns: {"hj_noten": {name: note15}, "mdl_noten_kurs": {name: {f"{hj_key}_mdl1": v, ...}}}
    """
    max_col = ws.max_column or 20
    note_col: Optional[int] = None
    for c in range(1, max_col + 1):
        v = ws.cell(4, c).value
        if isinstance(v, str) and v.strip().lower() == "note":
            note_col = c
            break

    hj_noten: dict = {}
    mdl_noten_kurs: dict = {}

    for r in range(6, ws.max_row + 1):
        raw_name = ws.cell(r, 1).value
        if not raw_name:
            continue
        name = str(raw_name).strip()
        if not name or name.startswith("#") or name.lower() in _SKIP_NAMES:
            continue

        # MDL1 from col D (4), MDL2 from col E (5)
        mdl1 = _note_from_cell(ws, r, 4)
        mdl2 = _note_from_cell(ws, r, 5)
        if mdl1 is not None or mdl2 is not None:
            mdl_noten_kurs.setdefault(name, {})
            if mdl1 is not None:
                mdl_noten_kurs[name][f"{hj_key}_mdl1"] = mdl1
            if mdl2 is not None:
                mdl_noten_kurs[name][f"{hj_key}_mdl2"] = mdl2

        # Zeugnisnote from note_col (two rows below "Note" header: row 6 not row 4+2)
        if note_col:
            note = _note_from_cell(ws, r, note_col)
            if note is not None:
                hj_noten[name] = note

    return {"hj_noten": hj_noten, "mdl_noten_kurs": mdl_noten_kurs}


# ── Oberstufe: full import orchestration ─────────────────────────────────────

def import_oberstufe_legacy_file(wb, selections: dict, students: list[str]) -> dict:
    """
    Import an Oberstufen-Kurs legacy file.
    Returns a complete gradebook dict with modus='kurs'.
    """
    lns: list[dict] = []
    mdl_noten_kurs: dict = {}
    hj_noten: dict = {}

    auto_slots = selections.get("gln_auto_slots", {})

    # ── GLN sheets ────────────────────────────────────────────────────────────
    for gname in selections.get("gln_sheets_selected", []):
        slot_info = auto_slots.get(gname, {})
        gln_slot = slot_info.get("slot", "GLN1")
        hj = slot_info.get("hj", "HJ1")
        ln = _read_oberstufe_gln(wb, gname, students, gln_slot, hj)
        if ln:
            lns.append(ln)

    # ── Zeugnisnoten sheets ───────────────────────────────────────────────────
    for zname in _ZEUGNIS_HJ_MAP:
        if zname not in wb.sheetnames:
            continue
        hj_key = _ZEUGNIS_HJ_MAP[zname]
        result = _read_zeugnisnoten(wb[zname], students, hj_key)
        # Merge hj_noten
        for name, note in result["hj_noten"].items():
            hj_noten.setdefault(name, {})[hj_key] = note
        # Merge mdl_noten_kurs
        for name, vals in result["mdl_noten_kurs"].items():
            mdl_noten_kurs.setdefault(name, {}).update(vals)

    # ── Build stammdaten ──────────────────────────────────────────────────────
    stammdaten = []
    for name in students:
        parts = name.split(",", 1)
        nachname = parts[0].strip()
        vorname = parts[1].strip() if len(parts) > 1 else ""
        stammdaten.append({
            "nachname": nachname,
            "vorname": vorname,
            "notizen": "",
            "status": S.SD_STATUS_AKTIV,
            "austritt": "",
        })

    # ── Auto-detect students who left early ──────────────────────────────────
    # Check which students appear in each Zeugnisnoten sheet (col A, row 6+).
    # If a student is missing from a later sheet but appeared in an earlier one,
    # they are marked as ausgeschieden after their last HJ.
    _hj_order = ["HJ1", "HJ2", "HJ3", "HJ4"]
    hj_students_present: dict[str, set] = {}
    for zname, hj_key in _ZEUGNIS_HJ_MAP.items():
        if zname not in wb.sheetnames:
            continue
        ws_z = wb[zname]
        present: set[str] = set()
        for r in range(6, ws_z.max_row + 1):
            v = ws_z.cell(r, 1).value
            if not v:
                continue
            n = str(v).strip()
            if n and not n.startswith("#") and n.lower() not in _SKIP_NAMES:
                present.add(n)
        if present:
            hj_students_present[hj_key] = present

    if len(hj_students_present) >= 2:
        max_hj_idx = max(
            _hj_order.index(hj) for hj in hj_students_present
        )
        for sd in stammdaten:
            full_name = f"{sd['nachname']}, {sd['vorname']}" if sd['vorname'] else sd['nachname']
            last_hj_idx = -1
            for idx, hj in enumerate(_hj_order):
                if full_name in hj_students_present.get(hj, set()):
                    last_hj_idx = idx
            if 0 <= last_hj_idx < max_hj_idx:
                sd["status"] = S.SD_STATUS_AUSGESCHIEDEN
                sd["abgang_nach_hj"] = _hj_order[last_hj_idx]


    return {
        "klasse":         selections.get("klasse", ""),
        "fach":           selections.get("fach", ""),
        "schuljahr":      selections.get("schuljahr", ""),
        "schuljahr_bis":  selections.get("schuljahr_bis", ""),
        "modus":          "kurs",
        "kurs_typ":       "GK",
        "kurs_stunden":   4,
        "kurs_gewichtung": {"hj_gln_pct": 70.0, "hj_mdl_pct": 30.0},
        "stammdaten":     stammdaten,
        "leistungsnachweise": lns,
        "mdl_noten_kurs": mdl_noten_kurs,
        "hj_noten":       hj_noten,
        "mdl_noten":      {},
        "sl_noten_actual": {},
        "schuljahr_noten_actual": {},
        "uebersicht_hj1": None,
        "uebersicht_hj2": None,
        "uebersicht_jahr": None,
        "sl_gewichtung":  None,
    }


# ── public: execute import ────────────────────────────────────────────────────

def import_legacy_file(file_bytes: bytes, password: Optional[str], selections: dict) -> dict:
    """
    Execute the import based on user selections from the wizard.
    Returns a dict compatible with the standard gradebook session format.
    """
    wb = _open_wb(file_bytes, password)

    # ── Delegate to Oberstufe import if flagged ───────────────────────────────
    if selections.get("is_oberstufe"):
        students = _get_students_oberstufe(wb) or _get_students(wb)
        return import_oberstufe_legacy_file(wb, selections, students)

    students = _get_students(wb)

    lns: list[dict] = []
    mdl_noten: dict = {}
    sl_noten_actual: dict = {}
    hj_noten: dict = {}
    schuljahr_noten_actual: dict = {}

    # ── KLN imports ───────────────────────────────────────────────────────────
    for sel in selections.get("kln_imports", []):
        ln = _read_kln_entry(wb, sel, students)
        if ln:
            lns.append(ln)

    # ── GLN imports ───────────────────────────────────────────────────────────
    for sel in selections.get("gln_imports", []):
        ln = _read_gln_entry(wb, sel, students)
        if ln:
            lns.append(ln)

    # ── bis_sl controls how much of the Noten sheets to import ─────────────
    bis_sl = selections.get("bis_sl", "SL4")  # "SL1" | "SL2" | "SL3" | "SL4"

    if bis_sl in ("SL1", "SL2", "SL3", "SL4") and "Noten" in wb.sheetnames:
        _read_noten_hj1(wb["Noten"], students, mdl_noten, sl_noten_actual, hj_noten, bis_sl)

    if bis_sl in ("SL3", "SL4") and "Noten (2)" in wb.sheetnames:
        _read_noten_hj2(wb["Noten (2)"], students, mdl_noten, sl_noten_actual, schuljahr_noten_actual, bis_sl)

    # ── determine inactive students (in Noten but not in Noten (2)) ──────────
    ausgeschieden_names: set = set()
    if "Noten (2)" in wb.sheetnames:
        ws2 = wb["Noten (2)"]
        noten2_names: set = set()
        for r2 in range(4, ws2.max_row + 1):
            raw = ws2.cell(r2, 1).value
            if not raw:
                continue
            n2 = str(raw).strip()
            if n2 and not n2.startswith("#") and n2.lower() not in _SKIP_NAMES:
                noten2_names.add(n2)
        if noten2_names:
            ausgeschieden_names = set(students) - noten2_names

    # ── build stammdaten ──────────────────────────────────────────────────────
    stammdaten = []
    for name in students:
        parts = name.split(",", 1)
        nachname = parts[0].strip()
        vorname = parts[1].strip() if len(parts) > 1 else ""
        status = S.SD_STATUS_AUSGESCHIEDEN if name in ausgeschieden_names else S.SD_STATUS_AKTIV
        stammdaten.append({"nachname": nachname, "vorname": vorname, "notizen": "",
                           "status": status, "austritt": ""})

    return {
        "klasse": selections.get("klasse", ""),
        "fach": selections.get("fach", ""),
        "schuljahr": selections.get("schuljahr", ""),
        "stammdaten": stammdaten,
        "leistungsnachweise": lns,
        "mdl_noten": mdl_noten,
        "sl_noten_actual": sl_noten_actual,
        "hj_noten": hj_noten,
        "schuljahr_noten_actual": schuljahr_noten_actual,
        "uebersicht_hj1": None,
        "uebersicht_hj2": None,
        "uebersicht_jahr": None,
        "sl_gewichtung": None,
    }


# ── KLN entry reader ──────────────────────────────────────────────────────────

def _read_kln_entry(wb: Workbook, sel: dict, students: list[str]) -> Optional[dict]:
    sheet_name: str = sel["sheet"]
    col: int = sel["col"]           # 1-based column for the HÜ points
    hue_name: str = sel["name"]
    max_pts: int = sel["max_pts"]
    hj: str = sel["hj"]
    sl: str = sel["sl"]

    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]

    # Unique internal name: e.g. "KLN1-HÜ1"
    ln_name = f"{sheet_name.replace(' ', '')}-{hue_name}"

    schueler = []
    for r in range(4, ws.max_row + 1):
        raw_name = ws.cell(r, 1).value
        if not raw_name:
            continue
        name = str(raw_name).strip()
        if not name or name.startswith("#") or name.lower() in _SKIP_NAMES:
            continue

        pts_raw = ws.cell(r, col).value
        try:
            pts = round(float(pts_raw), 2) if pts_raw is not None else None
        except (ValueError, TypeError):
            pts = None

        schueler.append(
            {
                "name": name,
                "punkte": [pts],
                "note_15": None,
                "note_6": None,
                "ignoriert": False,
            }
        )

    return {
        "sheet_name": f"LN_{ln_name}",
        "name": ln_name,
        "ln_typ": "KLN",
        "hj": hj,
        "sl_zuordnung": sl,
        "aufgaben": [{"label": "Gesamt", "afb": "", "max_punkte": max_pts}],
        "aufgaben_tree": [],
        "schueler": schueler,
    }


# ── GLN entry reader ──────────────────────────────────────────────────────────

def _read_gln_entry(wb: Workbook, sel: dict, students: list[str]) -> Optional[dict]:
    sheet_name: str = sel["sheet"]
    hj: str = sel["hj"]
    sl: Optional[str] = sel.get("sl")

    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]

    # Find "Gesamt" column index in row 2
    max_col = ws.max_column or 25
    gesamt_col: Optional[int] = None
    for c in range(1, max_col + 1):
        v = ws.cell(2, c).value
        if isinstance(v, str) and v.strip() == "Gesamt":
            gesamt_col = c
            break

    # Max points from row 3 col O (or Gesamt col)
    gesamt_max = 100
    if gesamt_col:
        raw = ws.cell(3, gesamt_col).value
        try:
            gesamt_max = round(float(raw)) if raw else 100
        except (ValueError, TypeError):
            gesamt_max = 100

    ln_name = sheet_name.replace(" ", "")
    schueler = []
    for r in range(4, ws.max_row + 1):
        raw_name = ws.cell(r, 1).value
        if not raw_name:
            continue
        name = str(raw_name).strip()
        if not name or name.startswith("#") or name.lower() in _SKIP_NAMES:
            continue

        # Note(15) after Gesamt col: Gesamt+1
        note15: Optional[int] = None
        note6: Optional[int] = None
        if gesamt_col:
            note15 = _int_or_none(ws.cell(r, gesamt_col + 1).value)
            note6 = _int_or_none(ws.cell(r, gesamt_col + 2).value)

        # punkte: try Gesamt column (often None for cross-sheet refs)
        pts: Optional[float] = None
        if gesamt_col:
            pts_raw = ws.cell(r, gesamt_col).value
            try:
                pts = round(float(pts_raw), 2) if pts_raw is not None else None
            except (ValueError, TypeError):
                pts = None

        schueler.append(
            {
                "name": name,
                "punkte": [pts],
                "note_15": note15,
                "note_6": note6,
                "ignoriert": False,
            }
        )

    return {
        "sheet_name": f"LN_{ln_name}",
        "name": ln_name,
        "ln_typ": "GLN",
        "hj": hj,
        "sl_zuordnung": sl,
        "aufgaben": [{"label": "Gesamt", "afb": "", "max_punkte": gesamt_max}],
        "aufgaben_tree": [],
        "schueler": schueler,
    }


# ── Noten sheet readers ───────────────────────────────────────────────────────

def _note_from_cell(ws, row: int, col: int) -> Optional[int]:
    v = ws.cell(row, col).value
    if v is None:
        return None
    try:
        f = float(v)
        if 0 <= f <= 15:
            return round(f)
        return None
    except (ValueError, TypeError):
        return None


def _read_noten_hj1(
    ws,
    students: list[str],
    mdl_noten: dict,
    sl_noten_actual: dict,
    hj_noten: dict,
    bis_sl: str = "SL2",
) -> None:
    """
    Extract from "Noten" sheet (HJ1):
      Col B (2)  = MDL 1  → mdl_noten[name]['SL1']
      Col C (3)  = MDL 2  → mdl_noten[name]['SL2']  (only if bis_sl != 'SL1')
      Col I (9)  = Note (HJ1 Zeugnisnote)            (only if bis_sl != 'SL1')
      Col O (15) = SL 1 end
      Col S (19) = SL2 end                           (only if bis_sl != 'SL1')
    """
    for r in range(4, ws.max_row + 1):
        raw_name = ws.cell(r, 1).value
        if not raw_name:
            continue
        name = str(raw_name).strip()
        if not name or name.startswith("#") or name.lower() in _SKIP_NAMES:
            continue

        mdl1 = _note_from_cell(ws, r, 2)
        sl1_end = _note_from_cell(ws, r, 15)

        if mdl1 is not None:
            mdl_noten.setdefault(name, {})["SL1"] = mdl1
        if sl1_end is not None:
            sl_noten_actual.setdefault(name, {})["SL1"] = sl1_end

        if bis_sl != "SL1":
            mdl2 = _note_from_cell(ws, r, 3)
            hj1_note = _note_from_cell(ws, r, 9)
            sl2_end = _note_from_cell(ws, r, 19)
            if mdl2 is not None:
                mdl_noten.setdefault(name, {})["SL2"] = mdl2
            if hj1_note is not None:
                hj_noten.setdefault(name, {})["HJ1"] = hj1_note
            if sl2_end is not None:
                sl_noten_actual.setdefault(name, {})["SL2"] = sl2_end


def _read_noten_hj2(
    ws,
    students: list[str],
    mdl_noten: dict,
    sl_noten_actual: dict,
    schuljahr_noten_actual: dict,
    bis_sl: str = "SL4",
) -> None:
    """
    Extract from "Noten (2)" sheet (full year):
      Col B (2)  = MDL3  → mdl_noten[name]['SL3']
      Col C (3)  = MDL4  → mdl_noten[name]['SL4']  (only if bis_sl == 'SL4')
      Col I (9)  = Ganzjahresnote                   (only if bis_sl == 'SL4')
      Col P (16) = SL 3 end
      Col T (20) = SL4 end                          (only if bis_sl == 'SL4')
    """
    for r in range(4, ws.max_row + 1):
        raw_name = ws.cell(r, 1).value
        if not raw_name:
            continue
        name = str(raw_name).strip()
        if not name or name.startswith("#") or name.lower() in _SKIP_NAMES:
            continue

        mdl3 = _note_from_cell(ws, r, 2)
        sl3_end = _note_from_cell(ws, r, 16)

        if mdl3 is not None:
            mdl_noten.setdefault(name, {})["SL3"] = mdl3
        if sl3_end is not None:
            sl_noten_actual.setdefault(name, {})["SL3"] = sl3_end

        if bis_sl == "SL4":
            mdl4 = _note_from_cell(ws, r, 3)
            gj_note = _note_from_cell(ws, r, 9)
            sl4_end = _note_from_cell(ws, r, 20)
            if mdl4 is not None:
                mdl_noten.setdefault(name, {})["SL4"] = mdl4
            if gj_note is not None:
                schuljahr_noten_actual[name] = gj_note
            if sl4_end is not None:
                sl_noten_actual.setdefault(name, {})["SL4"] = sl4_end
