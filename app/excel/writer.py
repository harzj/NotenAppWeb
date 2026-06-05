"""
Build a grade xlsx workbook from in-memory data and optionally encrypt it.
Formulas are written as formula strings so Excel recalculates on open.
"""
from __future__ import annotations

import io
from typing import Any

import msoffcrypto
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.excel import schema as S


# ── Colours ───────────────────────────────────────────────────────────────────
COLOR_HEADER_BG = "1F4E79"   # dark blue
COLOR_HEADER_FG = "FFFFFF"
COLOR_AFB_BG    = "BDD7EE"   # light blue
COLOR_MAX_BG    = "DDEBF7"
COLOR_NOTE_BG   = "E2EFDA"   # light green
COLOR_ALT_ROW   = "F2F2F2"


def build_gradebook(data: dict, password: str | None = None) -> bytes:
    """
    Build a complete xlsx from *data* (same structure as reader output).
    If *password* is given the file is encrypted with Excel workbook protection.
    Returns raw bytes suitable for a file download.
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    klasse = data.get("klasse", "")
    fach = data.get("fach", "")
    schuljahr = data.get("schuljahr", "")
    stammdaten = data.get("stammdaten", [])
    # Map "Nachname, Vorname" → row number in Stammdaten sheet (for formula references)
    name_to_sd_row: dict[str, int] = {
        f"{s['nachname']}, {s['vorname']}": S.SD_DATA_START_ROW + i
        for i, s in enumerate(stammdaten)
    }
    _write_stammdaten(wb, stammdaten, klasse=klasse, fach=fach, schuljahr=schuljahr)
    for ln in data.get("leistungsnachweise", []):
        _write_ln_sheet(wb, ln, name_to_sd_row=name_to_sd_row)
    _write_uebersicht(wb, data, S.SHEET_UEBERSICHT_HJ1, "HJ1")
    _write_uebersicht(wb, data, S.SHEET_UEBERSICHT_HJ2, "HJ2")
    if data.get("modus") == "kurs":
        _write_uebersicht(wb, data, S.SHEET_UEBERSICHT_HJ3, "HJ3")
        _write_uebersicht(wb, data, S.SHEET_UEBERSICHT_HJ4, "HJ4")
    _write_uebersicht(wb, data, S.SHEET_UEBERSICHT_JAHR, "Jahr")
    _write_notentabelle(wb)
    _write_noten_zusatz(wb, data, stammdaten)
    _write_einstellungen(wb, data)

    raw = io.BytesIO()
    wb.save(raw)
    raw.seek(0)

    if password:
        return _encrypt(raw, password)
    return raw.getvalue()


# ── Stammdaten ────────────────────────────────────────────────────────────────

def _write_stammdaten(wb: Workbook, students: list[dict], klasse: str = "", fach: str = "", schuljahr: str = "") -> None:
    ws = wb.create_sheet(S.SHEET_STAMMDATEN)
    # Row 1: class name + fach + schuljahr metadata
    lbl = ws.cell(S.SD_INFO_ROW, S.SD_CLASS_NAME_LABEL_COL, S.SD_CLASS_NAME_LABEL)
    lbl.font = Font(bold=True, size=11)
    val = ws.cell(S.SD_INFO_ROW, S.SD_CLASS_NAME_VALUE_COL, klasse)
    val.font = Font(bold=True, size=12, color="1F4E79")
    fach_lbl = ws.cell(S.SD_INFO_ROW, S.SD_FACH_LABEL_COL, S.SD_FACH_LABEL)
    fach_lbl.font = Font(bold=True, size=11)
    fach_val = ws.cell(S.SD_INFO_ROW, S.SD_FACH_VALUE_COL, fach)
    fach_val.font = Font(bold=True, size=12, color="1F4E79")
    sj_lbl = ws.cell(S.SD_INFO_ROW, S.SD_SJ_LABEL_COL, S.SD_SJ_LABEL)
    sj_lbl.font = Font(bold=True, size=11)
    sj_val = ws.cell(S.SD_INFO_ROW, S.SD_SJ_VALUE_COL, schuljahr)
    sj_val.font = Font(bold=True, size=12, color="1F4E79")
    # Row 2: headers
    headers = ["Nachname", "Vorname", "Status", "Austrittsdatum", "Abgang_nach_HJ",
               "Aufnahme_ab_HJ", "Aufnahme_Vorh_HJ1", "Aufnahme_Vorh_HJ2", "Aufnahme_Vorh_HJ3"]
    _write_header_row(ws, S.SD_HEADER_ROW, headers)
    for i, s in enumerate(students, start=S.SD_DATA_START_ROW):
        ws.cell(i, S.SD_COL_NACHNAME, s.get("nachname", ""))
        ws.cell(i, S.SD_COL_VORNAME,  s.get("vorname",  ""))
        ws.cell(i, S.SD_COL_STATUS,   s.get("status",   S.SD_STATUS_AKTIV))
        ws.cell(i, S.SD_COL_AUSTRITT, s.get("austritt", ""))
        ws.cell(i, S.SD_COL_ABGANG_HJ, s.get("abgang_nach_hj", ""))
        ws.cell(i, S.SD_COL_AUFNAHME_HJ, s.get("aufnahme_ab_hj") or "")
        prev = s.get("aufnahme_vorherige_noten") or {}
        ws.cell(i, S.SD_COL_AUFNAHME_PREV_HJ1, prev.get("HJ1") if prev.get("HJ1") is not None else "")
        ws.cell(i, S.SD_COL_AUFNAHME_PREV_HJ2, prev.get("HJ2") if prev.get("HJ2") is not None else "")
        ws.cell(i, S.SD_COL_AUFNAHME_PREV_HJ3, prev.get("HJ3") if prev.get("HJ3") is not None else "")
    _autofit(ws)


# ── Leistungsnachweis ─────────────────────────────────────────────────────────

def _write_ln_sheet(wb: Workbook, ln: dict, name_to_sd_row: dict | None = None) -> None:
    ws = wb.create_sheet(ln["sheet_name"])
    aufgaben: list[dict] = ln.get("aufgaben", [])
    schueler: list[dict] = ln.get("schueler", [])

    n_tasks = len(aufgaben)
    col_gesamt    = S.LN_COL_TASKS_START + n_tasks          # 1-based
    col_note15    = col_gesamt + 1
    col_note6     = col_gesamt + 2
    col_ignoriert = col_gesamt + 3

    # ── Row 1: LN metadata (v2 format marker) ──
    meta_style = Font(bold=True, color="FFFFFF")
    meta_fill  = PatternFill("solid", fgColor="4F4F4F")
    for col, label, val in [
        (S.LN_META_TYP_COL, S.LN_META_TYP_LABEL, ln.get("ln_typ", "")),
        (S.LN_META_TYP_VAL, None, ln.get("ln_typ", "")),
        (S.LN_META_HJ_COL,  S.LN_META_HJ_LABEL,  None),
        (S.LN_META_HJ_VAL,  None, ln.get("hj") or ""),
        (S.LN_META_SL_COL,  S.LN_META_SL_LABEL,  None),
        (S.LN_META_SL_VAL,  None, ln.get("sl_zuordnung") or ""),
        (S.LN_META_GSLOT_COL, S.LN_META_GSLOT_LABEL, None),
        (S.LN_META_GSLOT_VAL, None, ln.get("gln_slot") or ""),
        (S.LN_META_NT_COL,    S.LN_META_NT_LABEL,    None),
        (S.LN_META_NT_VAL,    None, ln.get("nachtermin_von") or ""),
        (S.LN_META_RUNDEN_COL, S.LN_META_RUNDEN_LABEL, None),
        (S.LN_META_RUNDEN_VAL, None, "1" if ln.get("noten_runden", True) else "0"),
    ]:
        c = ws.cell(S.LN_ROW_META, col)
        c.value = label if label is not None else val
        c.fill  = meta_fill
        c.font  = Font(bold=True, color="FFFFFF")
    # Overwrite value cells with actual values
    ws.cell(S.LN_ROW_META, S.LN_META_TYP_VAL).value = ln.get("ln_typ", "")
    ws.cell(S.LN_ROW_META, S.LN_META_HJ_VAL).value  = ln.get("hj") or ""
    ws.cell(S.LN_ROW_META, S.LN_META_SL_VAL).value  = ln.get("sl_zuordnung") or ""
    ws.cell(S.LN_ROW_META, S.LN_META_NT_VAL).value  = ln.get("nachtermin_von") or ""
    ws.cell(S.LN_ROW_META, S.LN_META_RUNDEN_VAL).value = "1" if ln.get("noten_runden", True) else "0"

    # ── Row 2: headers ──
    ws.cell(S.LN_ROW_HEADER, S.LN_COL_NAME, S.LN_HEADER_NAME)
    for t_idx, task in enumerate(aufgaben):
        ws.cell(S.LN_ROW_HEADER, S.LN_COL_TASKS_START + t_idx, task["label"])
    is_abt = ln.get("ln_typ") == S.LN_TYP_ABT
    ws.cell(S.LN_ROW_HEADER, col_gesamt,    S.LN_HEADER_GESAMT)
    ws.cell(S.LN_ROW_HEADER, col_note15,    S.LN_HEADER_NOTE_15)
    ws.cell(S.LN_ROW_HEADER, col_note6,     S.LN_HEADER_NOTE_6)
    ws.cell(S.LN_ROW_HEADER, col_ignoriert, S.LN_HEADER_KUERZEL if is_abt else S.LN_HEADER_IGNORIERT)
    _style_header_row(ws, S.LN_ROW_HEADER, col_ignoriert)

    # ── Row 3: Anforderungsbereich ──
    ws.cell(S.LN_ROW_AFB, S.LN_COL_NAME, S.LN_HEADER_AFB_LABEL)
    for t_idx, task in enumerate(aufgaben):
        ws.cell(S.LN_ROW_AFB, S.LN_COL_TASKS_START + t_idx, task.get("afb", ""))
    _style_row_bg(ws, S.LN_ROW_AFB, col_ignoriert, COLOR_AFB_BG)

    # ── Row 4: Max. Punkte ──
    ws.cell(S.LN_ROW_MAX, S.LN_COL_NAME, S.LN_HEADER_MAX_LABEL)
    task_start_letter = get_column_letter(S.LN_COL_TASKS_START)
    task_end_letter   = get_column_letter(S.LN_COL_TASKS_START + n_tasks - 1)
    gesamt_letter     = get_column_letter(col_gesamt)

    for t_idx, task in enumerate(aufgaben):
        ws.cell(S.LN_ROW_MAX, S.LN_COL_TASKS_START + t_idx, task.get("max_punkte") or 0)

    if n_tasks > 0:
        ws.cell(S.LN_ROW_MAX, col_gesamt,
                f"=SUM({task_start_letter}{S.LN_ROW_MAX}:{task_end_letter}{S.LN_ROW_MAX})")
    _style_row_bg(ws, S.LN_ROW_MAX, col_ignoriert, COLOR_MAX_BG)

    # ── Rows 5+: Students ──
    nt_sheet = S.SHEET_NOTENTABELLE
    for s_idx, s in enumerate(schueler):
        row = S.LN_DATA_START_ROW + s_idx
        name = s["name"]
        sd_row = name_to_sd_row.get(name) if name_to_sd_row else None
        if sd_row is not None:
            sd = S.SHEET_STAMMDATEN
            ws.cell(row, S.LN_COL_NAME,
                    f'={sd}!A{sd_row}&", "&{sd}!B{sd_row}')
        else:
            ws.cell(row, S.LN_COL_NAME, name)

        punkte = s.get("punkte", [])
        for t_idx in range(n_tasks):
            val = punkte[t_idx] if t_idx < len(punkte) else None
            ws.cell(row, S.LN_COL_TASKS_START + t_idx, val)

        row_letter = str(row)
        if n_tasks > 0:
            ws.cell(row, col_gesamt,
                    f"=SUM({task_start_letter}{row_letter}:{task_end_letter}{row_letter})")

        if n_tasks > 0:
            ws.cell(row, col_note15,
                _note15_formula(gesamt_letter, row_letter,
                                gesamt_letter, str(S.LN_ROW_MAX), nt_sheet))
            ws.cell(row, col_note6,
                _note6_formula(get_column_letter(col_note15), row_letter))

        # Kürzel (ABT) or Ignoriert flag (non-ABT)
        if is_abt:
            ws.cell(row, col_ignoriert, s.get("kuerzel") or None)
        else:
            ws.cell(row, col_ignoriert, 1 if s.get("ignoriert") else None)

        if s_idx % 2 == 1:
            _style_row_bg(ws, row, col_ignoriert, COLOR_ALT_ROW)
        for c in (col_note15, col_note6):
            ws.cell(row, c).fill = PatternFill("solid", fgColor=COLOR_NOTE_BG)

    _autofit(ws)


def _note15_formula(gesamt_col: str, row: str, max_col: str, max_row: str, nt_sheet: str = "") -> str:
    """
    Self-contained nested-IF mapping achieved/max → note 0-15.
    Does not require the Notentabelle sheet (works in all Excel versions).
    nt_sheet kept for API compatibility.
    """
    achieved = f"{gesamt_col}{row}"
    maximum  = f"{max_col}{max_row}"
    pct      = f"({achieved}/{maximum})"
    # Build nested IF from inside out.
    # GRADE_SCALE is descending (0.95→15 … 0.20→1 … 0.0→0).
    # Reverse-iterate, skipping the last (0.0, 0) entry (the catch-all "else 0").
    result = "0"
    for threshold, note in reversed(S.GRADE_SCALE[:-1]):
        result = f"IF({pct}>={threshold},{note},{result})"
    return f"=IF({maximum}=0,0,{result})"


def _note6_formula(note15_col: str, row: str) -> str:
    """Map note 0-15 to note 1-6 using nested IFs."""
    c = f"{note15_col}{row}"
    return (
        f"=IF({c}>=13,1,"
        f"IF({c}>=10,2,"
        f"IF({c}>=7,3,"
        f"IF({c}>=4,4,"
        f"IF({c}>=1,5,6)))))"
    )


# ── Übersicht sheets ──────────────────────────────────────────────────────────

def _write_uebersicht(wb: Workbook, data: dict, sheet_name: str, hj_key: str) -> None:
    key_map = {
        S.SHEET_UEBERSICHT_HJ1: "uebersicht_hj1",
        S.SHEET_UEBERSICHT_HJ2: "uebersicht_hj2",
        S.SHEET_UEBERSICHT_JAHR: "uebersicht_jahr",
    }
    uebersicht = data.get(key_map.get(sheet_name, ""), None)
    ws = wb.create_sheet(sheet_name)

    # Always rebuild class overview sheets from current in-memory data so that
    # manually edited notes (e.g. Verhalten/Mitarbeit) are reflected in export.
    if sheet_name in (S.SHEET_UEBERSICHT_HJ1, S.SHEET_UEBERSICHT_HJ2, S.SHEET_UEBERSICHT_JAHR):
        _build_uebersicht_from_ln(wb, ws, data, sheet_name)
        return

    if not uebersicht:
        # Build from LN data
        _build_uebersicht_from_ln(wb, ws, data, sheet_name)
        return

    cols = uebersicht.get("columns", [])
    _write_header_row(ws, 1, cols)
    for r_idx, row in enumerate(uebersicht.get("rows", []), start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, val)
    _autofit(ws)


def _build_uebersicht_from_ln(wb: Workbook, ws, data: dict, sheet_name: str) -> None:
    """Auto-build an overview sheet referencing LN sheets."""
    all_lns: list[dict] = data.get("leistungsnachweise", [])

    def _ln_in_sheet(ln: dict) -> bool:
        # Ignore Nachtermin child sheets in overview export.
        if ln.get("nachtermin_von"):
            return False
        hj = (ln.get("hj") or "").strip()
        if sheet_name == S.SHEET_UEBERSICHT_HJ1:
            return hj == "HJ1"
        if sheet_name == S.SHEET_UEBERSICHT_HJ2:
            return hj == "HJ2"
        if sheet_name == S.SHEET_UEBERSICHT_HJ3:
            return hj == "HJ3"
        if sheet_name == S.SHEET_UEBERSICHT_HJ4:
            return hj == "HJ4"
        if sheet_name == S.SHEET_UEBERSICHT_JAHR:
            # Jahr sheet should include all regular HJ entries.
            return hj in ("HJ1", "HJ2", "HJ3", "HJ4")
        return True

    lns: list[dict] = [ln for ln in all_lns if _ln_in_sheet(ln)]
    students: list[dict] = data.get("stammdaten", [])
    verhalten_noten: dict = data.get("verhalten_noten") or {}
    mitarbeit_noten: dict = data.get("mitarbeit_noten") or {}
    with_social_cols = sheet_name in (
        S.SHEET_UEBERSICHT_HJ1,
        S.SHEET_UEBERSICHT_HJ2,
        S.SHEET_UEBERSICHT_JAHR,
    )

    headers = ["Schüler"] + [ln["name"] for ln in lns] + ["Ø (gewichtet)", "Zeugnisnote"]
    if with_social_cols:
        headers += ["Verhalten", "Mitarbeit"]
    _write_header_row(ws, 1, headers)

    def _pick_hj_val(notes: dict, name: str, key: str):
        if key == "HJ1":
            return notes.get(name, {}).get("HJ1")
        if key == "HJ2":
            return notes.get(name, {}).get("HJ2")
        # Jahr: prefer HJ2, fallback HJ1
        if key == "Jahr":
            v2 = notes.get(name, {}).get("HJ2")
            v1 = notes.get(name, {}).get("HJ1")
            return v2 if v2 is not None else v1
        return None

    for s_idx, student in enumerate(students):
        if student.get("status") == S.SD_STATUS_AUSGESCHIEDEN:
            continue
        row = s_idx + 2
        full_name = f"{student['nachname']}, {student['vorname']}"
        sd_row_num = S.SD_DATA_START_ROW + s_idx
        ws.cell(row, 1, f'={S.SHEET_STAMMDATEN}!A{sd_row_num}&", "&{S.SHEET_STAMMDATEN}!B{sd_row_num}')

        for ln_idx, ln in enumerate(lns):
            col = ln_idx + 2
            ln_ws_name = ln["sheet_name"]
            # Find student row in LN sheet
            ln_student_row = _find_student_row_in_ln(ln, full_name)
            if ln_student_row is not None:
                n_tasks = len(ln.get("aufgaben", []))
                note15_col = get_column_letter(S.LN_COL_TASKS_START + n_tasks + 1)
                ws.cell(row, col, f"='{ln_ws_name}'!{note15_col}{ln_student_row}")
            else:
                ws.cell(row, col, "")

        # Weighted average (simple mean here; weights can be extended)
        if lns:
            avg_start = get_column_letter(2)
            avg_end   = get_column_letter(1 + len(lns))
            avg_col   = get_column_letter(len(lns) + 2)
            ws.cell(row, len(lns) + 2, f"=AVERAGE({avg_start}{row}:{avg_end}{row})")
            ws.cell(row, len(lns) + 3, f"=ROUND({avg_col}{row},0)")

        if with_social_cols:
            hj_key = "HJ1" if sheet_name == S.SHEET_UEBERSICHT_HJ1 else (
                "HJ2" if sheet_name == S.SHEET_UEBERSICHT_HJ2 else "Jahr"
            )
            ws.cell(row, len(lns) + 4, _pick_hj_val(verhalten_noten, full_name, hj_key))
            ws.cell(row, len(lns) + 5, _pick_hj_val(mitarbeit_noten, full_name, hj_key))

    _autofit(ws)


def _find_student_row_in_ln(ln: dict, full_name: str) -> int | None:
    for s_idx, s in enumerate(ln.get("schueler", [])):
        if s.get("name", "").strip() == full_name.strip():
            return S.LN_DATA_START_ROW + s_idx
    return None


# ── Notentabelle (hidden reference sheet) ────────────────────────────────────

def _write_noten_zusatz(wb: Workbook, data: dict, stammdaten: list[dict]) -> None:
    """Write per-student MDL, SL-actual, HJ-actual and SJ-actual notes to a hidden sheet."""
    ws = wb.create_sheet(S.SHEET_NOTEN_ZUSATZ)
    ws.sheet_state = "hidden"

    _write_header_row(ws, S.NZ_HEADER_ROW, S.NZ_HEADERS)

    mdl_noten        = data.get("mdl_noten") or {}
    sl_noten_actual  = data.get("sl_noten_actual") or {}
    hj_noten         = data.get("hj_noten") or {}
    sj_noten_actual  = data.get("schuljahr_noten_actual") or {}
    mdl_noten_kurs   = data.get("mdl_noten_kurs") or {}
    verhalten_noten  = data.get("verhalten_noten") or {}
    mitarbeit_noten  = data.get("mitarbeit_noten") or {}

    for i, s in enumerate(stammdaten, start=S.NZ_DATA_START):
        name = f"{s['nachname']}, {s['vorname']}"
        ws.cell(i, S.NZ_COL_NAME, name)
        ws.cell(i, S.NZ_COL_MDL_SL1,    mdl_noten.get(name, {}).get("SL1"))
        ws.cell(i, S.NZ_COL_MDL_SL2,    mdl_noten.get(name, {}).get("SL2"))
        ws.cell(i, S.NZ_COL_MDL_SL3,    mdl_noten.get(name, {}).get("SL3"))
        ws.cell(i, S.NZ_COL_MDL_SL4,    mdl_noten.get(name, {}).get("SL4"))
        ws.cell(i, S.NZ_COL_SL_ACT_SL1, sl_noten_actual.get(name, {}).get("SL1"))
        ws.cell(i, S.NZ_COL_SL_ACT_SL2, sl_noten_actual.get(name, {}).get("SL2"))
        ws.cell(i, S.NZ_COL_SL_ACT_SL3, sl_noten_actual.get(name, {}).get("SL3"))
        ws.cell(i, S.NZ_COL_SL_ACT_SL4, sl_noten_actual.get(name, {}).get("SL4"))
        ws.cell(i, S.NZ_COL_HJ_ACT_HJ1, hj_noten.get(name, {}).get("HJ1"))
        ws.cell(i, S.NZ_COL_HJ_ACT_HJ2, hj_noten.get(name, {}).get("HJ2"))
        ws.cell(i, S.NZ_COL_SJ_ACT,     sj_noten_actual.get(name))
        # Kurs extra columns
        mdl_kurs = mdl_noten_kurs.get(name, {})
        ws.cell(i, S.NZ_COL_KURS_MDL_HJ1_1, mdl_kurs.get("HJ1_mdl1"))
        ws.cell(i, S.NZ_COL_KURS_MDL_HJ1_2, mdl_kurs.get("HJ1_mdl2"))
        ws.cell(i, S.NZ_COL_KURS_MDL_HJ2_1, mdl_kurs.get("HJ2_mdl1"))
        ws.cell(i, S.NZ_COL_KURS_MDL_HJ2_2, mdl_kurs.get("HJ2_mdl2"))
        ws.cell(i, S.NZ_COL_KURS_MDL_HJ3_1, mdl_kurs.get("HJ3_mdl1"))
        ws.cell(i, S.NZ_COL_KURS_MDL_HJ3_2, mdl_kurs.get("HJ3_mdl2"))
        ws.cell(i, S.NZ_COL_KURS_MDL_HJ4_1, mdl_kurs.get("HJ4_mdl1"))
        ws.cell(i, S.NZ_COL_KURS_MDL_HJ4_2, mdl_kurs.get("HJ4_mdl2"))
        ws.cell(i, S.NZ_COL_KURS_HJ_ACT_HJ3, hj_noten.get(name, {}).get("HJ3"))
        ws.cell(i, S.NZ_COL_KURS_HJ_ACT_HJ4, hj_noten.get(name, {}).get("HJ4"))
        ws.cell(i, S.NZ_COL_VERH_HJ1, verhalten_noten.get(name, {}).get("HJ1"))
        ws.cell(i, S.NZ_COL_VERH_HJ2, verhalten_noten.get(name, {}).get("HJ2"))
        ws.cell(i, S.NZ_COL_MIT_HJ1, mitarbeit_noten.get(name, {}).get("HJ1"))
        ws.cell(i, S.NZ_COL_MIT_HJ2, mitarbeit_noten.get(name, {}).get("HJ2"))


def _write_einstellungen(wb: Workbook, data: dict) -> None:
    """Write sl_gewichtung and Kurs settings as key-value pairs to a hidden sheet."""
    ws = wb.create_sheet(S.SHEET_EINSTELLUNGEN)
    ws.sheet_state = "hidden"
    _write_header_row(ws, S.ES_HEADER_ROW, ["Einstellung", "Wert"])
    gw = data.get("sl_gewichtung") or {}
    kgw = data.get("kurs_gewichtung") or {}
    row = S.ES_DATA_START
    for key in S.ES_GEWICHTUNG_KEYS:
        ws.cell(row, S.ES_COL_KEY,   key)
        ws.cell(row, S.ES_COL_VALUE, gw.get(key))
        row += 1
    # Kurs-specific settings
    kurs_vals = {
        "modus":       data.get("modus", "klasse"),
        "kurs_typ":    data.get("kurs_typ", ""),
        "kurs_stunden": data.get("kurs_stunden", ""),
        "kurs_gln_pct": kgw.get("hj_gln_pct", ""),
        "kurs_mdl_pct": kgw.get("hj_mdl_pct", ""),
    }
    for key in S.ES_KURS_KEYS:
        ws.cell(row, S.ES_COL_KEY,   key)
        ws.cell(row, S.ES_COL_VALUE, kurs_vals.get(key))
        row += 1


def _write_notentabelle(wb: Workbook) -> None:
    ws = wb.create_sheet(S.SHEET_NOTENTABELLE)
    ws.sheet_state = "hidden"
    ws.cell(S.NT_HEADER_ROW, S.NT_COL_PERCENT, "Untergrenze (%)")
    ws.cell(S.NT_HEADER_ROW, S.NT_COL_NOTE_15, "Note (0-15)")
    for i, (pct, note) in enumerate(S.GRADE_SCALE, start=S.NT_DATA_START_ROW):
        ws.cell(i, S.NT_COL_PERCENT, pct)
        ws.cell(i, S.NT_COL_NOTE_15, note)


# ── Encryption ────────────────────────────────────────────────────────────────

def _encrypt(raw: io.BytesIO, password: str) -> bytes:
    encrypted = io.BytesIO()
    office_file = msoffcrypto.OfficeFile(raw)
    office_file.encrypt(password, encrypted)
    return encrypted.getvalue()


# ── Styling helpers ───────────────────────────────────────────────────────────

def _write_header_row(ws, row: int, headers: list) -> None:
    for col, val in enumerate(headers, start=1):
        cell = ws.cell(row, col, val)
        cell.font = Font(bold=True, color=COLOR_HEADER_FG)
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30


def _style_header_row(ws, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row, col)
        cell.font = Font(bold=True, color=COLOR_HEADER_FG)
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_row_bg(ws, row: int, max_col: int, color: str) -> None:
    fill = PatternFill("solid", fgColor=color)
    for col in range(1, max_col + 1):
        ws.cell(row, col).fill = fill


def _autofit(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value) if cell.value is not None else ""
                if not val.startswith("="):
                    max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 40)
